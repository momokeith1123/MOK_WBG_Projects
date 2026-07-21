#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Generic FSMP SQL Report Generator
=================================

Reads a SQL file from disk,
executes it against FSMP,
exports results to formatted Excel.
"""

import os
import datetime
import pandas as pd
import oracledb
import openpyxl

from openpyxl.styles import (
    PatternFill,
    Font,
    Border,
    Side,
    Alignment
)

from openpyxl.utils import get_column_letter

import config


# ============================================================
# ORACLE CLIENT INITIALIZATION
# ============================================================

oracle_client_dir = config.ORACLE_CLIENT_DIR

oracle_network_admin = os.path.abspath(
    os.path.join(
        oracle_client_dir,
        "..",
        "network",
        "admin"
    )
)

os.environ["TNS_ADMIN"] = oracle_network_admin

try:
    oracledb.init_oracle_client(
        lib_dir=oracle_client_dir,
        config_dir=oracle_network_admin
    )
except oracledb.ProgrammingError:
    pass


# ============================================================
# USER INPUT
# ============================================================

SQL_FILE = r"C:\Users\mkeita9\Projects\MOK_WBG_Projects\SQL_Queries\trade_holding_query.sql"

ASOFDATE = datetime.date(2026, 7, 21)

REPORT_NAME = (
    os.path.splitext(
        os.path.basename(SQL_FILE)
    )[0]
)

OUTPUT_FILE = os.path.join(
    config.SAVE_PATH,
    f"{REPORT_NAME}_{ASOFDATE:%Y%m%d}.xlsx"
)


# ============================================================
# LOAD SQL
# ============================================================

print("Loading SQL file...")

with open(SQL_FILE, "r", encoding="utf-8") as f:
    SQL = f.read()

print(f"Loaded: {SQL_FILE}")


# ============================================================
# CONNECT
# ============================================================

print("Connecting to FSMP...")

conn = oracledb.connect(
    user=config.PERF_DB_USER,
    password=config.PERF_DB_PASSWORD,
    dsn=config.PERF_DB_DSN
)

print(
    f"Connected to {config.PERF_DB_DSN}"
)


# ============================================================
# EXECUTE SQL
# ============================================================

print(
    f"Running report for {ASOFDATE}"
)

with conn.cursor() as cursor:

    cursor.execute(
        SQL,
        {
            "asofdate": ASOFDATE,
            "asofDate": ASOFDATE
        }
    )

    rows = cursor.fetchall()

    columns = [
        c[0]
        for c in cursor.description
    ]

conn.close()

df = pd.DataFrame(
    rows,
    columns=columns
)

print(
    f"{len(df):,} rows returned"
)


# ============================================================
# SAVE TO EXCEL
# ============================================================

with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="xlsxwriter"
) as writer:

    df.to_excel(
        writer,
        sheet_name="Report",
        index=False
    )

print(
    "Excel file written."
)


# ============================================================
# FORMAT EXCEL
# ============================================================

print(
    "Applying formatting..."
)

wb = openpyxl.load_workbook(
    OUTPUT_FILE
)

ws = wb.active

HEADER_FILL = PatternFill(
    start_color="1F3864",
    end_color="1F3864",
    fill_type="solid"
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF"
)

ROW_LIGHT = PatternFill(
    start_color="FFFFFF",
    end_color="FFFFFF",
    fill_type="solid"
)

ROW_DARK = PatternFill(
    start_color="F2F5F9",
    end_color="F2F5F9",
    fill_type="solid"
)

thin = Side(
    style="thin",
    color="BDC3C7"
)

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

# HEADER

for cell in ws[1]:

    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = border

    cell.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

# ROWS

for row in range(2, ws.max_row + 1):

    fill = (
        ROW_DARK
        if row % 2 == 0
        else ROW_LIGHT
    )

    for col in range(1, ws.max_column + 1):

        cell = ws.cell(
            row=row,
            column=col
        )

        cell.fill = fill
        cell.border = border

# AUTOFIT

for col in ws.columns:

    max_len = 0

    col_letter = col[0].column_letter

    for cell in col:

        try:

            max_len = max(
                max_len,
                len(str(cell.value))
            )

        except:
            pass

    ws.column_dimensions[
        col_letter
    ].width = min(
        max_len + 3,
        40
    )

# DATE FIELDS

for col in ws.columns:

    header = str(
        col[0].value
    ).upper()

    if (
        "DATE" in header
        or header.endswith("_DT")
    ):

        for cell in col[1:]:

            cell.number_format = "MM/DD/YYYY"

# FREEZE

ws.freeze_panes = "A2"

# FILTER

ws.auto_filter.ref = ws.dimensions

wb.save(
    OUTPUT_FILE
)

print()
print("=" * 60)
print("REPORT COMPLETED")
print("=" * 60)
print(f"Output file:")
print(OUTPUT_FILE)
print("=" * 60)