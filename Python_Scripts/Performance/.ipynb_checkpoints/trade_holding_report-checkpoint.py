#!/usr/bin/env python
# coding: utf-8
# fmt: off
r"""
Trade Holding Detail Report
============================
Reads the SQL query from an external .sql file, connects to the FSMP
performance database using credentials from the central config.py,
runs the query for a user-defined date range, and writes the result
to a professionally formatted Excel file.

SQL file  : C:\Users\mkeita9\Projects\MOK_WBG_Projects\SQL_Queries\trade_holding_query.sql
Output    : TradeHoldingDetail_<SDATE>_to_<EDATE>.xlsx  (saved to config.SAVE_PATH)
Config    : C:\Users\mkeita9\Projects\MOK_WBG_Projects\Python_Scripts\Utils\config\config.py
"""


# ---------------------------------------------------------------------------
# 1. IMPORTS
# ---------------------------------------------------------------------------
import os
import sys
import datetime

import oracledb                          # replaces cx_Oracle
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment


# ---------------------------------------------------------------------------
# 2. LOAD CENTRAL CONFIG
# ---------------------------------------------------------------------------
CONFIG_DIR = r'C:\Users\mkeita9\Projects\MOK_WBG_Projects\Python_Scripts\Utils\config'


if CONFIG_DIR not in sys.path:
    sys.path.insert(0, CONFIG_DIR)

try:
    import config
    print("=" * 70)
    print("CENTRAL CONFIG LOADED")
    print(f"  Config folder : {CONFIG_DIR}")
    print(f"  User          : {config.PERF_DB_USER}")
    print(f"  Save path     : {config.SAVE_PATH}")
    print("=" * 70)
except ModuleNotFoundError:
    raise SystemExit(
        "\n[ERROR] config.py not found.\n"
        f"Please place config.py in:\n  {CONFIG_DIR}\n"
        "and fill in your credentials and paths before running."
    )


# ---------------------------------------------------------------------------
# 3. PATHS
# ---------------------------------------------------------------------------
# Folder containing the .sql file
SQL_FOLDER = r'C:\Users\mkeita9\Projects\MOK_WBG_Projects\SQL_Queries'

# Name of the SQL file to run
SQL_FILE   = 'trade_holding_query.sql'

# Output folder — overrides config.SAVE_PATH for this specific script
OUTPUT_PATH = r'C:\Users\mkeita9\Projects\MOK_WBG_Projects\Python_Scripts\Performance'

# ---------------------------------------------------------------------------
# 4. DATE RANGE — edit these two lines to change the reporting period
# ---------------------------------------------------------------------------
SDATE = datetime.date(2026, 7, 20)   # start date (inclusive)
EDATE = datetime.date(2026, 7, 20)   # end date   (inclusive)

# Oracle expects dates as strings in 'DD-Mon-YYYY' format
sdate_str = SDATE.strftime('%d-%b-%Y').upper()   # e.g. '17-JUL-2026'
edate_str = EDATE.strftime('%d-%b-%Y').upper()   # e.g. '17-JUL-2026'

print(f"\nReporting period  : {sdate_str}  →  {edate_str}")


# ---------------------------------------------------------------------------
# 5. READ SQL FROM FILE
# ---------------------------------------------------------------------------
sql_path = os.path.join(SQL_FOLDER, SQL_FILE)

if not os.path.exists(sql_path):
    raise SystemExit(
        f"\n[ERROR] SQL file not found:\n  {sql_path}\n"
        "Please check that the file exists at the path above."
    )

with open(sql_path, 'r') as f:
    raw_sql = f.read()

# Strip comment-only lines (lines starting with --) so oracledb does not
# choke on the inline documentation at the top of the file.
sql_lines   = raw_sql.splitlines()
clean_lines = [ln for ln in sql_lines if not ln.strip().startswith('--')]
SQL         = '\n'.join(clean_lines)

print(f"SQL file loaded   : {sql_path}")


# ---------------------------------------------------------------------------
# 6. ORACLE CLIENT INITIALISATION
# ---------------------------------------------------------------------------
# oracledb runs in Thin mode by default — no Oracle Instant Client required.
# If your database requires Thick mode, set ORACLE_CLIENT_DIR in config.py.
# Otherwise set it to None or leave it out entirely.

try:
    oracle_client_dir = config.ORACLE_CLIENT_DIR
    if oracle_client_dir:
        # Thick mode — uses the local Oracle Instant Client
        oracledb.init_oracle_client(lib_dir=oracle_client_dir)
        print(f"Oracle mode       : Thick ({oracle_client_dir})")
    else:
        # Thin mode — no client installation needed
        print("Oracle mode       : Thin (no client required)")
except AttributeError:
    # ORACLE_CLIENT_DIR not defined in config.py — default to Thin mode
    print("Oracle mode       : Thin (ORACLE_CLIENT_DIR not set in config)")
except oracledb.ProgrammingError:
    # Client already initialised (safe to ignore in notebook/multi-script sessions)
    pass


# ---------------------------------------------------------------------------
# 7. DATABASE CONNECTION & QUERY EXECUTION
# ---------------------------------------------------------------------------
print(f"\nConnecting to     : {config.PERF_DB_USER}@{config.PERF_DB_DSN}")

connection = oracledb.connect(
    user=config.PERF_DB_USER,
    password=config.PERF_DB_PASSWORD,
    dsn=config.PERF_DB_DSN,
)

print("Connection        : OK")
print("Running query     : please wait, this may take a moment...\n")

with connection.cursor() as cursor:
    cursor.execute(SQL, sdate=sdate_str, edate=edate_str)
    rows    = cursor.fetchall()
    columns = [desc[0] for desc in cursor.description]

connection.close()

df = pd.DataFrame(rows, columns=columns)
print(f"Query complete    : {len(df):,} rows  ×  {len(df.columns)} columns")


# ---------------------------------------------------------------------------
# 8. COLUMN CLASSIFICATION
# ---------------------------------------------------------------------------
# These lists drive the number/date formatting applied in Section 10.

DATE_COLS = ['AS_OF_DATE', 'RUN_DT', 'MATURITY_DT']

AMOUNT_COLS = [
    'NOTIONAL_BEG_BOOK_CCY_AMT',
    'NOTIONAL_CHG_BOOK_CCY_AMT',
    'CURR_NTL_BOOK_CCY_AMT',
    'MARKET_VALUE_AMT',
    'MARKET_VALUE_BOOK_CCY_AMT',
    'RECV_MKT_VAL_AMT',
    'RECV_MKT_VAL_BOOK_CCY_AMT',
    'PENDING_CF_BOOK_CCY_AMT',
    'PENDING_CF_ADJ_AMT',
    'PENDING_CF_ADJ_BK_CCY_AMT',
    'ACCRUAL_BOOK_CCY_AMT',
    'CAPITAL_GL_BOOK_CCY_AMT',
    'CUM_CAPITAL_GL_AMT',
    'CURRENCY_GL_AMT',
    'DAILY_CF_BOOK_CCY_AMT',
    'FEE_BOOK_CCY_AMT',
    'INCOME_AMT',
    'INCOME_BOOK_CCY_AMT',
    'INCOME_DISCREPANCY_AMT',
    'RECEIVABLE_INT_BOOK_CCY_AMT',
]

INTEGER_COLS = ['VERSION_NBR']


# ---------------------------------------------------------------------------
# 9. WRITE RAW DATA TO INTERMEDIATE EXCEL FILE
# ---------------------------------------------------------------------------
dump_path = os.path.join(OUTPUT_PATH, '_dump_trade_holding.xlsx')

with pd.ExcelWriter(dump_path, engine='xlsxwriter') as writer:
    df.to_excel(writer, sheet_name='Trade Holding Detail', index=False)

print(f"\nRaw data written  : {dump_path}")


# ---------------------------------------------------------------------------
# 10. PROFESSIONAL FORMATTING
# ---------------------------------------------------------------------------

# ---- Corporate color palette ----------------------------------------------
HEADER_FILL = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Calibri')
ROW_LIGHT   = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
ROW_DARK    = PatternFill(start_color='F2F5F9', end_color='F2F5F9', fill_type='solid')

# ---- Column-group accent tints (applied to header cells only) -------------
ACCENT_ID   = 'BDD7EE'   # light blue   — position identifiers & dates
ACCENT_SEC  = 'D9E1F2'   # pale indigo  — security attributes
ACCENT_BOOK = 'E2EFDA'   # pale green   — book / trade attributes
ACCENT_NOTL = 'FFF2CC'   # pale amber   — notional amounts
ACCENT_MV   = 'FCE4D6'   # pale orange  — market values
ACCENT_CF   = 'EDEDED'   # light grey   — pending cash flows
ACCENT_PERF = 'EAD1DC'   # pale rose    — performance / P&L

# Map every column name to its accent color
ACCENT_MAP = {
    # Identifiers
    'AS_OF_DATE'               : ACCENT_ID,
    'RUN_DT'                   : ACCENT_ID,
    'VERSION_NBR'              : ACCENT_ID,
    'SECURITY_ID'              : ACCENT_ID,
    'TRADE_ID'                 : ACCENT_ID,
    # Security attributes
    'ASSET_TYPE_CODE'          : ACCENT_SEC,
    'NOTIONAL_CURRENCY_CODE'   : ACCENT_SEC,
    'MATURITY_DT'              : ACCENT_SEC,
    # Book / trade attributes
    'BOOK_CURRENCY_CODE'       : ACCENT_BOOK,
    'BOOK_CODE'                : ACCENT_BOOK,
    'ACCOUNT_CODE'             : ACCENT_BOOK,
    'PORTFOLIO_TYPE_CODE'      : ACCENT_BOOK,
    # Notional amounts
    'NOTIONAL_BEG_BOOK_CCY_AMT': ACCENT_NOTL,
    'NOTIONAL_CHG_BOOK_CCY_AMT': ACCENT_NOTL,
    'CURR_NTL_BOOK_CCY_AMT'    : ACCENT_NOTL,
    # Market values
    'MARKET_VALUE_AMT'           : ACCENT_MV,
    'MARKET_VALUE_BOOK_CCY_AMT'  : ACCENT_MV,
    'RECV_MKT_VAL_AMT'           : ACCENT_MV,
    'RECV_MKT_VAL_BOOK_CCY_AMT'  : ACCENT_MV,
    # Pending cash flows
    'PENDING_CF_BOOK_CCY_AMT'    : ACCENT_CF,
    'PENDING_CF_ADJ_AMT'         : ACCENT_CF,
    'PENDING_CF_ADJ_BK_CCY_AMT'  : ACCENT_CF,
    # Performance / P&L
    'ACCRUAL_BOOK_CCY_AMT'       : ACCENT_PERF,
    'CAPITAL_GL_BOOK_CCY_AMT'    : ACCENT_PERF,
    'CUM_CAPITAL_GL_AMT'         : ACCENT_PERF,
    'CURRENCY_GL_AMT'            : ACCENT_PERF,
    'DAILY_CF_BOOK_CCY_AMT'      : ACCENT_PERF,
    'FEE_BOOK_CCY_AMT'           : ACCENT_PERF,
    'INCOME_AMT'                 : ACCENT_PERF,
    'INCOME_BOOK_CCY_AMT'        : ACCENT_PERF,
    'INCOME_DISCREPANCY_AMT'     : ACCENT_PERF,
    'RECEIVABLE_INT_BOOK_CCY_AMT': ACCENT_PERF,
}

THIN_SIDE   = Side(style='thin', color='BDC3C7')
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE,
                     top=THIN_SIDE,  bottom=THIN_SIDE)


def col_letter_for(df, col_name):
    """Return the Excel column letter for a given DataFrame column name."""
    return get_column_letter(df.columns.get_loc(col_name) + 1)


def col_letters_for(df, col_names):
    """Return a list of Excel column letters for a list of column names."""
    return [col_letter_for(df, c) for c in col_names if c in df.columns]


# ---- Load workbook --------------------------------------------------------
print("Applying professional formatting...")

wb = openpyxl.load_workbook(dump_path)
ws = wb.active
max_row = ws.max_row
max_col = ws.max_column

# ---- Header row: navy fill, white bold, centered, bordered ----------------
ws.row_dimensions[1].height = 36
for cell in ws[1]:
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.border    = THIN_BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center',
                               wrap_text=True)

# ---- Apply column-group accent tints to header cells ----------------------
for col_idx in range(1, max_col + 1):
    col_name = str(ws.cell(row=1, column=col_idx).value or '').upper()
    if col_name in ACCENT_MAP:
        hex_col = ACCENT_MAP[col_name]
        cell    = ws.cell(row=1, column=col_idx)
        cell.fill = PatternFill(start_color=hex_col, end_color=hex_col,
                                fill_type='solid')
        cell.font = Font(bold=True, color='1F3864', size=11, name='Calibri')

# ---- Data rows: banding, borders, default alignment ----------------------
print(f"  Formatting {max_row - 1:,} data rows — please wait...")

for row_idx in range(2, max_row + 1):
    row_fill = ROW_DARK if row_idx % 2 == 0 else ROW_LIGHT
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill   = row_fill
        cell.border = THIN_BORDER
        cell.font   = Font(name='Calibri', size=10)
        if isinstance(cell.value, (int, float)):
            cell.alignment = Alignment(horizontal='right',  vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left',   vertical='center')

# ---- Date columns: MM/DD/YYYY, centered -----------------------------------
for col_ltr in col_letters_for(df, DATE_COLS):
    for row_idx in range(2, max_row + 1):
        cell = ws[f'{col_ltr}{row_idx}']
        cell.number_format = 'MM/DD/YYYY'
        cell.alignment     = Alignment(horizontal='center', vertical='center')

# ---- Amount columns: thousands separator, 2 decimals ---------------------
for col_ltr in col_letters_for(df, AMOUNT_COLS):
    for row_idx in range(2, max_row + 1):
        ws[f'{col_ltr}{row_idx}'].number_format = '#,##0.00'

# ---- Integer columns: no decimals ----------------------------------------
for col_ltr in col_letters_for(df, INTEGER_COLS):
    for row_idx in range(2, max_row + 1):
        ws[f'{col_ltr}{row_idx}'].number_format = '0'

# ---- Auto-fit column widths (sampled from header + first 50 data rows) ---
for col_idx in range(1, max_col + 1):
    col_ltr     = get_column_letter(col_idx)
    header_len  = len(str(ws.cell(row=1, column=col_idx).value or ''))
    sample_lens = [
        len(str(ws.cell(row=r, column=col_idx).value or ''))
        for r in range(2, min(max_row, 51))
    ]
    ws.column_dimensions[col_ltr].width = min(
        max(max([header_len] + sample_lens) + 2, 10), 40
    )

# ---- Freeze panes (keep header + first 3 ID columns visible) -------------
ws.freeze_panes = ws['D2']

# ---- Auto-filter ----------------------------------------------------------
ws.auto_filter.ref = ws.dimensions

# ---- Sheet tab color ------------------------------------------------------
ws.sheet_properties.tabColor = '1F3864'


# ---------------------------------------------------------------------------
# 11. SAVE FINAL OUTPUT FILE
# ---------------------------------------------------------------------------
output_filename = (
    f'TradeHoldingDetail_{SDATE.strftime("%Y%m%d")}'
    f'_to_{EDATE.strftime("%Y%m%d")}.xlsx'
)
output_path = os.path.join(OUTPUT_PATH, output_filename)

wb.save(output_path)

# Remove the intermediate dump file
if os.path.exists(dump_path):
    os.remove(dump_path)

print(f"\n{'=' * 70}")
print(f"OUTPUT SAVED")
print(f"  File   : {output_path}")
print(f"  Rows   : {len(df):,}")
print(f"  Columns: {len(df.columns)}")
print(f"{'=' * 70}")
print("Done ✅")
