#!/usr/bin/env python
# coding: utf-8
"""
Risk Sheet Generator — v4
=========================
Connects to two Oracle databases (FSMP and ORAMRP1), runs a series of SQL
queries, merges the resulting DataFrames, applies professional formatting,
and produces two Excel output files:

  1. NewRiskSheet_<YYYYMMDD>_Assumptions.xlsx        -> for traders
  2. NewRiskSheet_<YYYYMMDD>_Assumptions_update.xlsx -> for the risk team

All user-specific settings are read from config.py.

"""

# ---------------------------------------------------------------------------
# 1. IMPORTS
# ---------------------------------------------------------------------------
import os
import datetime

import oracledb
import pandas as pd
import numpy as np
from pandas.tseries.offsets import BDay

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import win32com.client as win32


# ---------------------------------------------------------------------------
# 2. LOAD CONFIGURATION
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# LOAD CONFIGURATION FROM CENTRAL LOCATION
# ---------------------------------------------------------------------------
import sys

# Path to the shared config folder — same for everyone on this machine layout
CONFIG_DIR = r'C:\Users\mkeita9\Projects\MOK_WBG_Projects\Python_Scripts\Utils\config'

if CONFIG_DIR not in sys.path:
    sys.path.insert(0, CONFIG_DIR)

try:
    import config
    print("=" * 70)
    print("CENTRAL CONFIG LOADED")
    print(f"  Config folder : {CONFIG_DIR}")
    print(f"  User          : {config.PERF_DB_USER}")
    print(f"  Save path     : {config.SAVE_PATH}")
    print("=" * 70)
except ModuleNotFoundError:
    raise SystemExit(
        "\n[ERROR] config.py not found.\n"
        f"Please place config.py in:\n  {CONFIG_DIR}\n"
        "and fill in your credentials and paths before running."
    )

# ---------------------------------------------------------------------------
# 3. ORACLE CLIENT INITIALIZATION
# ---------------------------------------------------------------------------
# Your machine resolves FSMP.worldbank.org through Oracle LDAP.
# tnsping confirmed that Oracle Client 19c is installed here:
# C:\WBG\oracle\product\19.0.0\client_1
#
# config.ORACLE_CLIENT_DIR should be:
# C:\WBG\oracle\product\19.0.0\client_1\bin

print("Initializing Oracle Client...")

oracle_client_dir = config.ORACLE_CLIENT_DIR

if oracle_client_dir:
    oracle_network_admin = os.path.abspath(
        os.path.join(oracle_client_dir, "..", "network", "admin")
    )

    if os.path.isdir(oracle_network_admin):
        os.environ["TNS_ADMIN"] = oracle_network_admin
        print(f"Oracle network admin : {oracle_network_admin}")
    else:
        print(
            f"Warning: Oracle network admin folder not found: "
            f"{oracle_network_admin}"
        )

    try:
        oracledb.init_oracle_client(
            lib_dir=oracle_client_dir,
            config_dir=oracle_network_admin if os.path.isdir(oracle_network_admin) else None,
        )
        print(f"Oracle Client loaded : {oracle_client_dir}")

    except oracledb.ProgrammingError as err:
        if "already initialized" in str(err).lower():
            print("Oracle Client already initialized.")
        else:
            raise
else:
    try:
        oracledb.init_oracle_client()
        print("Oracle Client initialized using system PATH.")
    except oracledb.ProgrammingError as err:
        if "already initialized" in str(err).lower():
            print("Oracle Client already initialized.")
        else:
            raise


# ---------------------------------------------------------------------------
# 4. LOAD PATHS
# ---------------------------------------------------------------------------
SQL_PATH = config.SQL_PATH
SAVE_PATH = config.SAVE_PATH

os.makedirs(SAVE_PATH, exist_ok=True)


# ---------------------------------------------------------------------------
# 5. SQL FILE REGISTRY
# ---------------------------------------------------------------------------
SQL_DICT = {
    "1. output.sql": {"sheet": "output", "tb1": [0, 0]},
    "2. cash balance.sql": {"sheet": "cash balance", "tb1": [0, 0]},
    "3. pending cash.sql": {"sheet": "pending cash", "tb1": [0, 0]},
    "4. futures margin.sql": {"sheet": "futures margin", "tb1": [0, 0]},
    "5. quantity.sql": {"sheet": "quantity", "tb1": [0, 0]},
    "6. OTC ratings.sql": {"sheet": "OTC ratings", "tb1": [0, 0]},
    "7. Risk rating SP.sql": {"sheet": "Risk rating SP", "tb1": [0, 0]},
    "8. SWAP(T) info.sql": {"sheet": "SWAP(T) info", "tb1": [0, 0]},
    "9. Assumptions.sql": {"sheet": "Assumptions", "tb1": [0, 0]},
    "10. fx.sql": {"sheet": "fx", "tb1": [0, 0]},
    "b1. BBG download.sql": {"sheet": "bbg", "tb1": [0, 0]},
}

print("SQL file registry loaded.")


# ---------------------------------------------------------------------------
# 6. DATABASE CONNECTIONS
# ---------------------------------------------------------------------------
print("Connecting to databases...")

perf_connection = oracledb.connect(
    user=config.PERF_DB_USER,
    password=config.PERF_DB_PASSWORD,
    dsn=config.PERF_DB_DSN,
)

risk_connection = oracledb.connect(
    user=config.RISK_DB_USER,
    password=config.RISK_DB_PASSWORD,
    dsn=config.RISK_DB_DSN,
)

print(f"oracledb version : {oracledb.__version__}")
print(f"PERF connection  : {config.PERF_DB_USER}@{config.PERF_DB_DSN}")
print(f"RISK connection  : {config.RISK_DB_USER}@{config.RISK_DB_DSN}")


# ---------------------------------------------------------------------------
# 7. DATE SETUP
# ---------------------------------------------------------------------------
location = config.OFFICE_LOCATION.upper()

if location == "SG_LONDON":
    asof = (datetime.datetime.today() - BDay(2)).date()

elif location == "DC":
    today = datetime.datetime.today()
    if today.weekday() in [5, 6]:
        asof = (today - BDay(1)).date()
    else:
        asof = today.date()

elif location == "MANUAL":
    asof = config.MANUAL_DATE

else:
    raise ValueError(
        f"Unknown OFFICE_LOCATION '{config.OFFICE_LOCATION}' in config.py. "
        "Valid values: 'SG_LONDON', 'DC', 'MANUAL'."
    )

asofdate = {"asofdate": asof}
asofdate_str = {"asofdate_str": asof.strftime("%Y%m%d")}

print(f"Office location  : {config.OFFICE_LOCATION}")
print(f"As-of date       : {asofdate}")
print(f"As-of date string: {asofdate_str}")


# ---------------------------------------------------------------------------
# 8. PRE-RUN CLEANUP
# ---------------------------------------------------------------------------
file_name = f'NewRiskSheet_{asof.strftime("%Y%m%d")}_Assumptions_update.xlsx'
file_path = os.path.join(SAVE_PATH, file_name)

if os.path.exists(file_path):
    os.remove(file_path)
    print(f"Deleted existing file : {file_path}")
else:
    print(f"No pre-existing file  : {file_path}")


# ---------------------------------------------------------------------------
# 9. HELPER — EXECUTE SQL FILE
# ---------------------------------------------------------------------------
def execute_scripts(filename, path, connection, params=None, sql_registry=None):
    """
    Read a .sql file, execute it, and return the result as a DataFrame.
    """
    full_path = os.path.join(path, filename)

    if not os.path.exists(full_path):
        print(f"  [{filename}] ERROR: SQL file not found at {full_path}")
        return None

    with open(full_path, "r", encoding="utf-8") as fd:
        sql_content = fd.read()

    sql_commands = [cmd.strip() for cmd in sql_content.split(";") if cmd.strip()]

    for idx, command in enumerate(sql_commands, start=1):
        try:
            with connection.cursor() as cursor:
                if params:
                    cursor.execute(command, params)
                else:
                    cursor.execute(command)

                result_rows = cursor.fetchall()

                if cursor.description is None:
                    print(f"  [{filename}] table {idx} — no result set")
                    return pd.DataFrame()

                columns = [desc[0] for desc in cursor.description]
                df = pd.DataFrame(result_rows, columns=columns)

                print(f"  [{filename}] table {idx} — shape {df.shape} — OK")
                return df

        except oracledb.Error as err:
            print(f"  [{filename}] table {idx} — ERROR: {err}")
            print(f"  Snippet: {command[:120]!r}")

    return None


# ---------------------------------------------------------------------------
# 10. EXECUTE ALL SQL QUERIES
# ---------------------------------------------------------------------------
dfs = {}

print("\n--- Running SQL queries ---")

for key in SQL_DICT:

    if key in (
        "1. output.sql",
        "2. cash balance.sql",
        "3. pending cash.sql",
        "4. futures margin.sql",
        "5. quantity.sql",
        "9. Assumptions.sql",
        "10. fx.sql",
    ):
        df = execute_scripts(key, SQL_PATH, perf_connection, asofdate, SQL_DICT)

    elif key in ("6. OTC ratings.sql", "7. Risk rating SP.sql"):
        df = execute_scripts(key, SQL_PATH, perf_connection, None, SQL_DICT)

    elif key == "8. SWAP(T) info.sql":
        df = execute_scripts(key, SQL_PATH, perf_connection, asofdate_str, SQL_DICT)

    elif key == "b1. BBG download.sql":
        df = execute_scripts(key, SQL_PATH, risk_connection, asofdate, SQL_DICT)

    else:
        df = None

    if df is None:
        raise RuntimeError(f"SQL query failed or returned None for file: {key}")

    dfs[key] = df

print("--- All queries complete ---\n")


# ---------------------------------------------------------------------------
# 11. DATA TRANSFORMATION — BUILD MERGED RISK SHEET
# ---------------------------------------------------------------------------
concat_keys = [
    "1. output.sql",
    "2. cash balance.sql",
    "3. pending cash.sql",
    "4. futures margin.sql",
]

columns_to_remove = [
    "PENDING_CF_ADJ_AMT",
    "PENDING_CF_ADJ_BK_CCY_AMT",
    "PENDING_CF_AMT",
    "PENDING_CF_BOOK_CCY_AMT",
    "TOTAL_INST_DV01",
]

for key in concat_keys:
    cols_present = [c for c in columns_to_remove if c in dfs[key].columns]
    if cols_present:
        dfs[key] = dfs[key].drop(columns=cols_present)

concat_df = pd.concat([dfs[key] for key in concat_keys], ignore_index=True)


def country_of_risk(row):
    coll_ctry = row.get("COLL_CTRY", "")
    issuer_cntry = row.get("ISSUER_CNTRY", "")

    if pd.isna(coll_ctry):
        coll_ctry = ""

    if pd.isna(issuer_cntry):
        issuer_cntry = ""

    coll_ctry = str(coll_ctry)
    issuer_cntry = str(issuer_cntry)

    if len(coll_ctry) >= 2 and coll_ctry != issuer_cntry:
        return coll_ctry

    return issuer_cntry


concat_rating = pd.concat(
    [dfs["6. OTC ratings.sql"], dfs["7. Risk rating SP.sql"]],
    ignore_index=True,
)

concat_rating["CNTRY_OF_RISK"] = concat_rating.apply(country_of_risk, axis=1)


def conditional_keep_string(security_id):
    if pd.isna(security_id):
        return security_id

    security_id = str(security_id)

    if security_id.startswith(("REPO", "FXFWD", "FXSPOT", "SWAP")):
        parts = security_id.split("_")
        if len(parts) > 2:
            return "_".join(parts[:2])

    return security_id


merged_df = concat_df.copy()
merged_df["temp"] = merged_df["SECURITY_ID"].apply(conditional_keep_string)

rating_cols = [
    "SECURITY_ID",
    "PARENT_CUST",
    "SP",
    "MOODY",
    "FITCH",
    "ISSUER_CUST",
    "ISSUER_CNTRY",
    "CNTRY_OF_RISK",
]

merged_df = pd.merge(
    merged_df,
    concat_rating[rating_cols],
    left_on="temp",
    right_on="SECURITY_ID",
    how="left",
)

merged_df["ISSUER_CNTRY_x"] = (
    merged_df["ISSUER_CNTRY_x"]
    .replace("", pd.NA)
    .fillna(merged_df["ISSUER_CNTRY_y"])
)

merged_df.rename(columns={"ISSUER_CNTRY_x": "ISSUER_CNTRY"}, inplace=True)

merged_df["CNTRY_OF_RISK_x"] = (
    merged_df["CNTRY_OF_RISK_x"]
    .replace("", pd.NA)
    .fillna(merged_df["CNTRY_OF_RISK_y"])
)

merged_df.rename(columns={"CNTRY_OF_RISK_x": "CNTRY_OF_RISK"}, inplace=True)
merged_df.rename(columns={"SECURITY_ID_x": "SECURITY_ID"}, inplace=True)

merged_df.drop(
    columns=["temp", "SECURITY_ID_y", "CNTRY_OF_RISK_y", "ISSUER_CNTRY_y"],
    inplace=True,
    errors="ignore",
)


def move_column_before(df, col_to_move, anchor_col):
    if col_to_move not in df.columns:
        print(f"Warning: column to move not found: {col_to_move}")
        return df

    if anchor_col not in df.columns:
        print(f"Warning: anchor column not found: {anchor_col}")
        return df

    cols = df.columns.tolist()
    cols.remove(col_to_move)
    cols.insert(cols.index(anchor_col), col_to_move)

    return df[cols]


for key in dfs:

    if key == "5. quantity.sql":
        merged_df = pd.merge(
            merged_df,
            dfs[key][["SEC", "NET_NTL", "BOOK_CODE"]],
            left_on=["SEC_TRADE_ID", "BOOK_CODE"],
            right_on=["SEC", "BOOK_CODE"],
            how="left",
        )

        merged_df.loc[
            merged_df["ASSET_TYPE_CODE"] == "PendingCASH",
            "NET_NTL",
        ] = ""

        merged_df.loc[
            merged_df["ASSET_TYPE_CODE"] == "PendingCASH",
            "SECURITY_ID",
        ] = "\u00A0"

    elif key == "8. SWAP(T) info.sql":
        merged_df = pd.merge(
            merged_df,
            dfs[key][["TRADEID", "EXPDATE", "EFFDATE", "MATDATE", "STRIKE", "PORR"]],
            left_on="SECURITY_ID",
            right_on="TRADEID",
            how="left",
        )

    elif key == "9. Assumptions.sql":
        filtered_df = dfs[key].drop(
            columns=["IDENTIFIERTYPE", "IDENTIFIER"],
            errors="ignore",
        )

        merged_df = pd.merge(
            merged_df,
            filtered_df,
            left_on="SEC_TRADE_ID",
            right_on="SECID",
            how="left",
        )

        merged_df.loc[
            merged_df["ASSET_TYPE_CODE"] == "PendingCASH",
            filtered_df.columns,
        ] = ""


merged_df.rename(
    columns={
        "PARENT_CUST": "PARENT",
        "SP": "S&P",
        "MOODY": "MOODYs",
        "ISSUER_CUST": "ISSUER/CUST",
        "EXPDATE": "SWOPT EXPDT",
        "EFFDATE": "SWAP/T EFFDT",
        "MATDATE": "SWAP/T MATDT",
        "STRIKE": "SWAP/T STRIKE",
        "PORR": "SWAP/T PayRec",
    },
    inplace=True,
)

merged_df.drop(columns=["TRADEID", "SECID"], inplace=True, errors="ignore")

merged_df = move_column_before(merged_df, "FIXFLOAT", "Manual Schedule?")
merged_df = move_column_before(merged_df, "EXTID", "Manual Schedule?")
merged_df = move_column_before(merged_df, "NET_NTL", "QUANTITY")

merged_df.drop(columns=["QUANTITY", "SEC"], inplace=True, errors="ignore")
merged_df.rename(columns={"NET_NTL": "QUANTITY"}, inplace=True)


merged_df = pd.merge(
    merged_df,
    dfs["b1. BBG download.sql"],
    left_on="OTH",
    right_on="INSTRUMENT_ID",
    how="left",
)

merged_df = move_column_before(merged_df, "SECURITY_DES", "MATURITY_DT")
merged_df = move_column_before(merged_df, "RTG_DBRS", "SWOPT EXPDT")
merged_df = move_column_before(merged_df, "MM_MDY_RTG_SHRT", "SWOPT EXPDT")

merged_df.drop(columns=["INSTRUMENT_ID"], inplace=True, errors="ignore")

merged_df.rename(
    columns={
        "SECURITY_DES": "BLG TICKER",
        "RTG_DBRS": "DBRS",
        "MM_MDY_RTG_SHRT": "ST Moodys",
    },
    inplace=True,
)


# ---------------------------------------------------------------------------
# 12. WRITE INTERMEDIATE TRADERS FILE
# ---------------------------------------------------------------------------
dump_path = os.path.join(SAVE_PATH, "dump.xlsx")

with pd.ExcelWriter(dump_path, engine="xlsxwriter") as writer:
    merged_df.to_excel(writer, sheet_name="RS with Assumptions Data", index=False)

print("Intermediate dump.xlsx written.")


# ---------------------------------------------------------------------------
# 13. PROFESSIONAL FORMATTING HELPERS
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
ROW_LIGHT = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ROW_DARK = PatternFill(start_color="F2F5F9", end_color="F2F5F9", fill_type="solid")

ACCENT_BBG = "9DC3E6"
ACCENT_RATING = "D9E1F2"
ACCENT_SWAP = "A9D18E"
ACCENT_ASSUMPT = "FFE699"

THIN_SIDE = Side(style="thin", color="BDC3C7")
THIN_BORDER = Border(
    left=THIN_SIDE,
    right=THIN_SIDE,
    top=THIN_SIDE,
    bottom=THIN_SIDE,
)


def _col_range(start_letter, end_letter):
    s = column_index_from_string(start_letter)
    e = column_index_from_string(end_letter)
    return [get_column_letter(i) for i in range(s, e + 1)]


def apply_professional_formatting(
    file_path,
    date_columns,
    amount_columns=None,
    price_columns=None,
    dv01_columns=None,
    accent_groups=None,
    freeze_cell="D2",
):
    amount_columns = amount_columns or []
    price_columns = price_columns or []
    dv01_columns = dv01_columns or []
    accent_groups = accent_groups or []

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    ws.row_dimensions[1].height = 32

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    for start_col, end_col, hex_color in accent_groups:
        tint = PatternFill(
            start_color=hex_color,
            end_color=hex_color,
            fill_type="solid",
        )

        for col_letter in _col_range(start_col, end_col):
            hdr = ws[f"{col_letter}1"]
            hdr.fill = tint
            hdr.font = Font(bold=True, color="1F3864", size=11, name="Calibri")

    for row_idx in range(2, max_row + 1):
        row_fill = ROW_DARK if row_idx % 2 == 0 else ROW_LIGHT

        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.font = Font(name="Calibri", size=10)

            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    for col_letter in date_columns:
        for row_idx in range(2, max_row + 1):
            cell = ws[f"{col_letter}{row_idx}"]
            cell.number_format = "MM/DD/YYYY"
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for col_letter in amount_columns:
        for row_idx in range(2, max_row + 1):
            ws[f"{col_letter}{row_idx}"].number_format = "#,##0.00"

    for col_letter in price_columns:
        for row_idx in range(2, max_row + 1):
            ws[f"{col_letter}{row_idx}"].number_format = "0.0000"

    for col_letter in dv01_columns:
        for row_idx in range(2, max_row + 1):
            ws[f"{col_letter}{row_idx}"].number_format = "0.00"

    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)

        header_len = len(str(ws.cell(row=1, column=col_idx).value or ""))

        sample_lens = [
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(2, min(max_row, 51) + 1)
        ]

        ws.column_dimensions[col_letter].width = min(
            max(max([header_len] + sample_lens) + 2, 10),
            40,
        )

    ws.freeze_panes = ws[freeze_cell]
    ws.auto_filter.ref = ws.dimensions

    wb.save(file_path)
    print(f"  Formatting applied -> {file_path}")


# ---------------------------------------------------------------------------
# 14. FORMAT AND SAVE TRADERS FILE
# ---------------------------------------------------------------------------
print("\n--- Formatting traders' file ---")

apply_professional_formatting(
    file_path=dump_path,
    date_columns=["A", "H", "CN", "CO", "CP"],
    amount_columns=[
        "W", "X", "Y",
        "AA",
        "AB", "AC", "AD",
        "AE", "AF",
        "AG",
        "AH", "AI",
        "AJ", "AK",
    ],
    price_columns=["Z", "AL", "AM", "AN", "AO"],
    dv01_columns=_col_range("AP", "CF"),
    accent_groups=[
        ("G", "G", ACCENT_BBG),
        ("CJ", "CM", ACCENT_RATING),
        ("CN", "CR", ACCENT_SWAP),
        ("CU", "DD", ACCENT_ASSUMPT),
    ],
    freeze_cell="D2",
)

traders_file = os.path.join(
    SAVE_PATH,
    f'NewRiskSheet_{asof.strftime("%Y%m%d")}_Assumptions.xlsx',
)

os.replace(dump_path, traders_file)

print(f"Traders' file saved -> {traders_file}")
print("File for traders generated.")


# ---------------------------------------------------------------------------
# 15. BUILD RISK TEAM VERSION
# ---------------------------------------------------------------------------
fx_df = dfs["10. fx.sql"]

risk_df = pd.merge(
    merged_df,
    fx_df[["CURRENCY_CODE", "FX_RATE"]],
    left_on="CURRENCY_CODE",
    right_on="CURRENCY_CODE",
    how="left",
)

risk_df = pd.merge(
    risk_df,
    fx_df[["CURRENCY_CODE", "FX_RATE"]],
    left_on="BOOK_CURRENCY_CODE",
    right_on="CURRENCY_CODE",
    how="left",
)

risk_df["CURR_NTL_USD"] = risk_df["CURR_NTL_AMT"] * risk_df["FX_RATE_x"]

risk_df["MARKET_VALUE_USD"] = (
    risk_df["MARKET_VALUE_AMT"] * risk_df["FX_RATE_x"]
)

risk_df["MARKET_VALUE_BOOK_CCY_AMT_update"] = (
    risk_df["MARKET_VALUE_USD"] / risk_df["FX_RATE_y"]
)

risk_df = move_column_before(risk_df, "CURR_NTL_USD", "PRICE")
risk_df = move_column_before(risk_df, "MARKET_VALUE_USD", "CAPITAL_GL_AMT")
risk_df = move_column_before(
    risk_df,
    "MARKET_VALUE_BOOK_CCY_AMT_update",
    "MARKET_VALUE_BOOK_CCY_AMT",
)

risk_df.drop(
    columns=[
        "MARKET_VALUE_BOOK_CCY_AMT",
        "FX_RATE_x",
        "CURRENCY_CODE_y",
        "FX_RATE_y",
    ],
    inplace=True,
    errors="ignore",
)

risk_df.rename(
    columns={
        "MARKET_VALUE_BOOK_CCY_AMT_update": "MARKET_VALUE_BOOK_CCY_AMT",
        "CURRENCY_CODE_x": "CURRENCY_CODE",
    },
    inplace=True,
)

dump_risk_path = os.path.join(SAVE_PATH, "dump_risk.xlsx")

with pd.ExcelWriter(dump_risk_path, engine="xlsxwriter") as writer:
    risk_df.to_excel(writer, sheet_name="output", index=False)

print("Intermediate dump_risk.xlsx written.")


# ---------------------------------------------------------------------------
# 16. FORMAT AND SAVE RISK TEAM FILE
# ---------------------------------------------------------------------------
print("\n--- Formatting risk team file ---")

apply_professional_formatting(
    file_path=dump_risk_path,
    date_columns=["A", "H", "CN", "CO", "CP"],
    amount_columns=[
        "W", "X", "Y",
        "AA",
        "AB", "AC", "AD",
        "AE", "AF",
        "AG",
        "AH", "AI",
        "AJ", "AK",
    ],
    price_columns=["Z", "AL", "AM", "AN", "AO"],
    dv01_columns=_col_range("AP", "CF"),
    accent_groups=[
        ("G", "G", ACCENT_BBG),
        ("CJ", "CM", ACCENT_RATING),
        ("CN", "CR", ACCENT_SWAP),
        ("CU", "DD", ACCENT_ASSUMPT),
    ],
    freeze_cell="D2",
)

risk_file = os.path.join(
    SAVE_PATH,
    f'NewRiskSheet_{asof.strftime("%Y%m%d")}_Assumptions_update.xlsx',
)

os.replace(dump_risk_path, risk_file)

print(f"Risk team file saved -> {risk_file}")


# ---------------------------------------------------------------------------
# 17. MAKE FILE COMPATIBLE WITH THE RISK DATABASE (Excel COM re-save)
# ---------------------------------------------------------------------------

print("\nFinalising risk team file via Excel COM — this may take a moment...")

excel  = None
wb_com = None

try:
    # Use Dispatch instead of gencache.EnsureDispatch — more reliable
    excel = win32.Dispatch('Excel.Application')
    excel.Visible          = False   # run Excel silently in the background
    excel.DisplayAlerts    = False   # suppress any pop-up warnings

    wb_com = excel.Workbooks.Open(risk_file)
    wb_com.Save()
    wb_com.Close(False)
    print("File for risk team generated ✅")

except Exception as e:
    print(f"[WARNING] Excel COM re-save failed: {e}")
    print("The file was saved by openpyxl and should still be usable.")
    print("If the risk database rejects it, open and re-save it manually in Excel.")

finally:
    # Always clean up COM objects to avoid orphaned Excel processes
    try:
        if wb_com is not None:
            wb_com.Close(False)
    except:
        pass
    try:
        if excel is not None:
            excel.Quit()
    except:
        pass


# ---------------------------------------------------------------------------
# 18. CLOSE DATABASE CONNECTIONS
# ---------------------------------------------------------------------------
try:
    perf_connection.close()
    risk_connection.close()
    print("Database connections closed.")
except Exception as e:
    print(f"Warning: issue while closing database connections: {e}")